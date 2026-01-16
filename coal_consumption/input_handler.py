import openpyxl
import os
from typing import Dict, List, Optional, Any, Tuple
from utils import (
    load_config, get_default_parameters, get_calculation_settings, get_output_settings,
    validate_numeric, validate_positive, extract_parameters_without_notes, safe_divide
)

class InputHandler:
    """
    数据输入处理类
    负责处理用户输入的数据和从Excel文件导入的数据
    """
    
    def __init__(self):
        # 加载配置和参数
        self.config = load_config()
        self.default_parameters = get_default_parameters()
        self.calculation_settings = get_calculation_settings()
        self.output_settings = get_output_settings()
    
    def get_default_parameters(self) -> Dict[str, Any]:
        """
        获取默认参数
        
        Returns:
            默认参数字典
        """
        return self.default_parameters.copy()
    
    def get_calculation_settings(self) -> Dict[str, Any]:
        """
        获取计算设置
        
        Returns:
            计算设置字典
        """
        return self.calculation_settings.copy()
    
    def get_output_settings(self) -> Dict[str, Any]:
        """
        获取输出设置
        
        Returns:
            输出设置字典
        """
        return self.output_settings.copy()
    
    def get_user_input(self, parameters: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        """
        获取用户输入的参数
        
        Args:
            parameters: 可选的参数字典，如果提供则使用这些参数
            
        Returns:
            完整的参数字典
        """
        if parameters is None:
            parameters = {}
        
        # 合并默认参数和用户输入参数
        input_params = self.default_parameters.copy()
        input_params.update(parameters)
        
        # 验证参数
        return self.validate_parameters(input_params)
    
    def load_from_excel(self, file_path: str, sheet_name: str = 'Sheet1') -> Dict[str, Any]:
        """
        从Excel文件加载参数
        
        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称
            
        Returns:
            参数字典
        """
        if not os.path.exists(file_path):
            print(f"错误: Excel文件不存在: {file_path}")
            return self.default_parameters.copy()
        
        try:
            # 加载Excel文件
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb[sheet_name]
            
            # 读取参数
            params = {}
            
            # 假设参数存储在A列和B列
            for row in range(1, ws.max_row + 1):
                key_cell = ws.cell(row=row, column=1)
                value_cell = ws.cell(row=row, column=2)
                
                if key_cell.value and value_cell.value:
                    params[key_cell.value] = value_cell.value
            
            # 合并默认参数
            input_params = self.default_parameters.copy()
            input_params.update(params)
            
            # 验证参数
            return self.validate_parameters(input_params)
            
        except Exception as e:
            print(f"从Excel加载参数时出错: {e}")
            return self.default_parameters.copy()
    
    def load_benchmark_and_monthly_data(self, file_path: str, sheet_name: str = '25年对比') -> Dict[str, Any]:
        """
        从Excel文件加载基准数据和各月份数据
        
        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称
            
        Returns:
            包含基准数据和各月份数据的字典
        """
        if not os.path.exists(file_path):
            print(f"错误: Excel文件不存在: {file_path}")
            return {
                'benchmark': {},
                'months': {}
            }
        
        try:
            # 加载Excel文件
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb[sheet_name]
            
            result = {
                'benchmark': {},
                'months': {}
            }
            
            # 影响因素映射 - 行号到参数名
            factor_mapping = {
                4: '海水温度',  # 行4
                5: '当地气温',  # 行5
                6: '煤种热值',  # 行6
                7: '电煤水分',  # 行7
                8: '灰分',      # 行8
                9: '发电量',     # 行9
                11: '机组负荷率', # 行11
                16: '供电基准煤耗', # 行16
                17: '2#机组发电量', # 行17
                19: '2#机组负荷率', # 行19
                22: '2#机组供电基准煤耗' # 行22
            }
            
            # 基准数据列（D列，第4列）
            BENCHMARK_COL = 4
            
            # 月份数据列映射
            month_columns = {
                '1月': {'data': 6, 'factor': 7},  # 1月数据在F列(6)，影响系数在G列(7)
                '2月': {'data': 8, 'factor': 9},  # 2月数据在H列(8)，影响系数在I列(9)
                # 可根据需要添加更多月份
            }
            
            # 读取基准数据
            for row_num, factor_name in factor_mapping.items():
                cell = ws.cell(row=row_num, column=BENCHMARK_COL)
                if cell.value is not None:
                    # 跳过单位行，只处理数值
                    if isinstance(cell.value, (int, float)):
                        result['benchmark'][factor_name] = cell.value
            
            # 读取各月份数据
            for month, columns in month_columns.items():
                result['months'][month] = {}
                for row_num, factor_name in factor_mapping.items():
                    data_cell = ws.cell(row=row_num, column=columns['data'])
                    if data_cell.value is not None and isinstance(data_cell.value, (int, float)):
                        result['months'][month][factor_name] = data_cell.value
                    
                    # 读取影响煤耗系数
                    factor_cell = ws.cell(row=row_num, column=columns['factor'])
                    if factor_cell.value is not None and isinstance(factor_cell.value, (int, float)):
                        result['months'][month][f'{factor_name}_影响系数'] = factor_cell.value
            
            return result
            
        except Exception as e:
            print(f"从Excel加载基准和月份数据时出错: {e}")
            return {
                'benchmark': {},
                'months': {}
            }
    
    def validate_parameters(self, parameters: Dict[str, Any]) -> Dict[str, Any]:
        """
        验证参数的有效性
        
        Args:
            parameters: 参数字典
            
        Returns:
            验证后的参数字典
        """
        validated_params = parameters.copy()
        
        # 验证负荷率 (0-100%)
        for key in ['base_load', 'actual_load', 'heating_load']:
            if key in validated_params:
                validated_params[key] = validate_numeric(validated_params[key], 0, 100)
        
        # 验证温度 (-40°C 到 80°C)
        for key in ['base_temperature', 'actual_temperature', 'base_sea_temperature']:
            if key in validated_params:
                validated_params[key] = validate_numeric(validated_params[key], -40, 80)
        
        # 验证压力 (0-50 MPa)
        for key in ['base_pressure', 'actual_pressure']:
            if key in validated_params:
                validated_params[key] = validate_numeric(validated_params[key], 0, 50)
        
        # 验证湿度 (0-100%)
        for key in ['base_humidity', 'actual_humidity']:
            if key in validated_params:
                validated_params[key] = validate_numeric(validated_params[key], 0, 100)
        
        # 验证效率 (0.5-1.0)
        for key in ['efficiency']:
            if key in validated_params:
                validated_params[key] = validate_numeric(validated_params[key], 0.5, 1.0)
        
        # 验证煤的发热量 (1000-8000 kcal/kg)
        for key in ['coal_calorific_value', 'actual_calorific_value']:
            if key in validated_params:
                validated_params[key] = validate_numeric(validated_params[key], 1000, 8000)
        
        # 验证发电量和供热量
        for key in ['electricity_output', 'heating_output']:
            if key in validated_params:
                validated_params[key] = validate_numeric(validated_params[key], 0)
        
        # 验证运行工况
        if 'operation_mode' in validated_params:
            valid_modes = ['normal', 'peak', 'low', 'heating']
            if validated_params['operation_mode'] not in valid_modes:
                validated_params['operation_mode'] = 'normal'
        
        return validated_params
    
    def get_parameters_for_calculation(self, user_input: Dict[str, Any]) -> Dict[str, Any]:
        """
        获取用于计算的参数
        
        Args:
            user_input: 用户输入的参数
            
        Returns:
            用于计算的参数字典
        """
        # 验证参数
        validated_params = self.validate_parameters(user_input)
        
        # 添加计算所需的衍生参数
        calculation_params = validated_params.copy()
        
        # 将煤的发热量从kcal/kg转换为kJ/kg (1 kcal = 4.1868 kJ)
        if 'coal_calorific_value' in calculation_params:
            calculation_params['coal_calorific_value_kj'] = calculation_params['coal_calorific_value'] * 4.1868
        
        if 'actual_calorific_value' in calculation_params:
            calculation_params['actual_calorific_value_kj'] = calculation_params['actual_calorific_value'] * 4.1868
        
        # 设置实际参数的默认值
        if 'actual_load' not in calculation_params:
            calculation_params['actual_load'] = calculation_params['base_load']
        
        if 'actual_temperature' not in calculation_params:
            calculation_params['actual_temperature'] = calculation_params['base_temperature']
        
        if 'actual_pressure' not in calculation_params:
            calculation_params['actual_pressure'] = calculation_params['base_pressure']
        
        if 'actual_humidity' not in calculation_params:
            calculation_params['actual_humidity'] = calculation_params['base_humidity']
        
        if 'actual_calorific_value' not in calculation_params:
            calculation_params['actual_calorific_value'] = calculation_params['coal_calorific_value']
            calculation_params['actual_calorific_value_kj'] = calculation_params.get('coal_calorific_value_kj', 20320.0)
        
        # 计算总负荷
        if 'total_load' not in calculation_params:
            calculation_params['total_load'] = calculation_params['base_load']
        
        return calculation_params
