import openpyxl
from openpyxl.utils import get_column_letter

# 加载Excel文件
wb = openpyxl.load_workbook('外部因素对煤耗的计算表1222.xlsx', data_only=True)

print('=== 查找关键参数 ===\n')

ws = wb['25年对比']

# 查找海水温度
print('1. 海水温度:')
for row in range(1, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        if cell.value and '海水' in str(cell.value):
            print(f'  {get_column_letter(col)}{row}: {cell.value}')
            # 查找右侧和下方的数值
            for offset in range(1, 6):
                right_cell = ws.cell(row=row, column=col+offset)
                if right_cell.value and isinstance(right_cell.value, (int, float)):
                    print(f'    -> 右侧{offset}列: {right_cell.value}')
                
                down_cell = ws.cell(row=row+offset, column=col)
                if down_cell.value and isinstance(down_cell.value, (int, float)):
                    print(f'    -> 下方{offset}行: {down_cell.value}')

print()

# 查找气温
print('2. 气温:')
for row in range(1, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        if cell.value and ('气温' in str(cell.value) or '温度' in str(cell.value)):
            print(f'  {get_column_letter(col)}{row}: {cell.value}')
            # 查找右侧和下方的数值
            for offset in range(1, 6):
                right_cell = ws.cell(row=row, column=col+offset)
                if right_cell.value and isinstance(right_cell.value, (int, float)):
                    print(f'    -> 右侧{offset}列: {right_cell.value}')
                
                down_cell = ws.cell(row=row+offset, column=col)
                if down_cell.value and isinstance(down_cell.value, (int, float)):
                    print(f'    -> 下方{offset}行: {down_cell.value}')

print()

# 查找热值、发热量
print('3. 热值/发热量:')
for row in range(1, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        if cell.value and ('热值' in str(cell.value) or '发热量' in str(cell.value) or '29306' in str(cell.value) or '29.306' in str(cell.value)):
            print(f'  {get_column_letter(col)}{row}: {cell.value}')
            # 查找右侧和下方的数值
            for offset in range(1, 6):
                right_cell = ws.cell(row=row, column=col+offset)
                if right_cell.value and isinstance(right_cell.value, (int, float)):
                    print(f'    -> 右侧{offset}列: {right_cell.value}')
                
                down_cell = ws.cell(row=row+offset, column=col)
                if down_cell.value and isinstance(down_cell.value, (int, float)):
                    print(f'    -> 下方{offset}行: {down_cell.value}')

print()

# 查找水分
print('4. 水分:')
for row in range(1, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        if cell.value and '水分' in str(cell.value):
            print(f'  {get_column_letter(col)}{row}: {cell.value}')
            # 查找右侧和下方的数值
            for offset in range(1, 6):
                right_cell = ws.cell(row=row, column=col+offset)
                if right_cell.value and isinstance(right_cell.value, (int, float)):
                    print(f'    -> 右侧{offset}列: {right_cell.value}')
                
                down_cell = ws.cell(row=row+offset, column=col)
                if down_cell.value and isinstance(down_cell.value, (int, float)):
                    print(f'    -> 下方{offset}行: {down_cell.value}')

print()

# 查找灰分
print('5. 灰分:')
for row in range(1, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        if cell.value and '灰分' in str(cell.value):
            print(f'  {get_column_letter(col)}{row}: {cell.value}')
            # 查找右侧和下方的数值
            for offset in range(1, 6):
                right_cell = ws.cell(row=row, column=col+offset)
                if right_cell.value and isinstance(right_cell.value, (int, float)):
                    print(f'    -> 右侧{offset}列: {right_cell.value}')
                
                down_cell = ws.cell(row=row+offset, column=col)
                if down_cell.value and isinstance(down_cell.value, (int, float)):
                    print(f'    -> 下方{offset}行: {down_cell.value}')

print()

# 查看前30行的数据，寻找可能的参数定义
print('=== 前30行数据 ===')
for row in range(1, min(31, ws.max_row + 1)):
    row_data = []
    for col in range(1, min(21, ws.max_column + 1)):  # 前20列
        cell = ws.cell(row=row, column=col)
        if cell.value is not None:
            cell_str = str(cell.value)
            if len(cell_str) > 15:
                cell_str = cell_str[:15] + "..."
            row_data.append(f'{get_column_letter(col)}:{cell_str}')
    if row_data:
        print(f'行{row}: {", ".join(row_data)}')
