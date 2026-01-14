import openpyxl
from openpyxl.utils import get_column_letter

# 加载Excel文件
wb = openpyxl.load_workbook('外部因素对煤耗的计算表1222.xlsx', data_only=True)

print('=== Excel文件前40行数据 ===\n')

ws = wb.active

# 查看前40行的所有数据
for row in range(1, min(41, ws.max_row + 1)):
    row_data = []
    for col in range(1, min(41, ws.max_column + 1)):  # 前40列
        cell = ws.cell(row=row, column=col)
        if cell.value is not None:
            cell_str = str(cell.value)
            if len(cell_str) > 20:
                cell_str = cell_str[:20] + "..."
            row_data.append(f'{get_column_letter(col)}:{cell_str}')
    if row_data:
        print(f'行{row}: {", ".join(row_data)}')

print('\n=== 查找包含"基准"、"29306"、"16.7"的单元格 ===\n')

# 查找特定关键字
keywords = ['基准', '29306', '16.7', '29.306', '29.3']
for keyword in keywords:
    print(f'关键字: {keyword}')
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value and keyword in str(cell.value):
                print(f'  {get_column_letter(col)}{row}: {cell.value}')
    print()
