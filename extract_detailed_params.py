import openpyxl
from openpyxl.utils import get_column_letter

# 加载Excel文件
wb = openpyxl.load_workbook('外部因素对煤耗的计算表1222.xlsx', data_only=True)

print('=== Excel中的关键参数值 ===\n')

# 查找包含关键数值的单元格
ws = wb.active

# 查找基准参数
print('基准参数:')
print('-' * 50)

# 查找温度相关
for row in range(1, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        if cell.value and isinstance(cell.value, (int, float)):
            # 检查周围是否有"温度"关键字
            for r_offset in range(-3, 4):
                for c_offset in range(-3, 4):
                    new_row = row + r_offset
                    new_col = col + c_offset
                    if new_row >= 1 and new_col >= 1 and new_row <= ws.max_row and new_col <= ws.max_column:
                        nearby_cell = ws.cell(row=new_row, column=new_col)
                        if nearby_cell.value and '温度' in str(nearby_cell.value):
                            print(f'{get_column_letter(col)}{row}: {cell.value} (附近有"温度")')
                            break
                else:
                    continue
                break

print()

# 查找压力相关
for row in range(1, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        if cell.value and isinstance(cell.value, (int, float)):
            # 检查周围是否有"压力"关键字
            for r_offset in range(-3, 4):
                for c_offset in range(-3, 4):
                    new_row = row + r_offset
                    new_col = col + c_offset
                    if new_row >= 1 and new_col >= 1 and new_row <= ws.max_row and new_col <= ws.max_column:
                        nearby_cell = ws.cell(row=new_row, column=new_col)
                        if nearby_cell.value and '压力' in str(nearby_cell.value):
                            print(f'{get_column_letter(col)}{row}: {cell.value} (附近有"压力")')
                            break
                else:
                    continue
                break

print()

# 查找负荷率相关
for row in range(1, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        if cell.value and isinstance(cell.value, (int, float)):
            # 检查周围是否有"负荷"关键字
            for r_offset in range(-3, 4):
                for c_offset in range(-3, 4):
                    new_row = row + r_offset
                    new_col = col + c_offset
                    if new_row >= 1 and new_col >= 1 and new_row <= ws.max_row and new_col <= ws.max_column:
                        nearby_cell = ws.cell(row=new_row, column=new_col)
                        if nearby_cell.value and '负荷' in str(nearby_cell.value):
                            print(f'{get_column_letter(col)}{row}: {cell.value} (附近有"负荷")')
                            break
                else:
                    continue
                break

print()

# 查找煤耗相关
for row in range(1, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        if cell.value and isinstance(cell.value, (int, float)):
            # 检查周围是否有"煤耗"关键字
            for r_offset in range(-3, 4):
                for c_offset in range(-3, 4):
                    new_row = row + r_offset
                    new_col = col + c_offset
                    if new_row >= 1 and new_col >= 1 and new_row <= ws.max_row and new_col <= ws.max_column:
                        nearby_cell = ws.cell(row=new_row, column=new_col)
                        if nearby_cell.value and '煤耗' in str(nearby_cell.value):
                            print(f'{get_column_letter(col)}{row}: {cell.value} (附近有"煤耗")')
                            break
                else:
                    continue
                break

print()

# 查找效率相关
for row in range(1, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        if cell.value and isinstance(cell.value, (int, float)):
            # 检查周围是否有"效率"关键字
            for r_offset in range(-3, 4):
                for c_offset in range(-3, 4):
                    new_row = row + r_offset
                    new_col = col + c_offset
                    if new_row >= 1 and new_col >= 1 and new_row <= ws.max_row and new_col <= ws.max_column:
                        nearby_cell = ws.cell(row=new_row, column=new_col)
                        if nearby_cell.value and '效率' in str(nearby_cell.value):
                            print(f'{get_column_letter(col)}{row}: {cell.value} (附近有"效率")')
                            break
                else:
                    continue
                break

print()

# 查找发热量相关
for row in range(1, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        if cell.value and isinstance(cell.value, (int, float)):
            # 检查周围是否有"发热量"或"29306"等关键字
            for r_offset in range(-3, 4):
                for c_offset in range(-3, 4):
                    new_row = row + r_offset
                    new_col = col + c_offset
                    if new_row >= 1 and new_col >= 1 and new_row <= ws.max_row and new_col <= ws.max_column:
                        nearby_cell = ws.cell(row=new_row, column=new_col)
                        if nearby_cell.value and ('发热量' in str(nearby_cell.value) or '29306' in str(nearby_cell.value)):
                            print(f'{get_column_letter(col)}{row}: {cell.value} (附近有"发热量"或"29306")')
                            break
                else:
                    continue
                break

print()

# 查找湿度相关
for row in range(1, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        if cell.value and isinstance(cell.value, (int, float)):
            # 检查周围是否有"湿度"关键字
            for r_offset in range(-3, 4):
                for c_offset in range(-3, 4):
                    new_row = row + r_offset
                    new_col = col + c_offset
                    if new_row >= 1 and new_col >= 1 and new_row <= ws.max_row and new_col <= ws.max_column:
                        nearby_cell = ws.cell(row=new_row, column=new_col)
                        if nearby_cell.value and '湿度' in str(nearby_cell.value):
                            print(f'{get_column_letter(col)}{row}: {cell.value} (附近有"湿度")')
                            break
                else:
                    continue
                break
