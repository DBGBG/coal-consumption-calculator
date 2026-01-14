import openpyxl
from openpyxl.utils import get_column_letter

# 加载Excel文件
wb = openpyxl.load_workbook('外部因素对煤耗的计算表1222.xlsx', data_only=True)
ws = wb.active

print('=== 详细数据分析 ===')
print()

# 查看公式单元格周围的数据
print('=== 公式单元格周围数据 ===')
formula_locations = [(43, 6), (44, 6), (45, 6), (46, 13)]  # (行, 列)

for row, col in formula_locations:
    print(f'\n=== 公式单元格周围 (行{row}, 列{get_column_letter(col)}) ===')
    
    # 查看上下5行，左右3列的数据
    for r in range(max(1, row-5), min(ws.max_row+1, row+6)):
        row_data = []
        for c in range(max(1, col-3), min(ws.max_column+1, col+4)):
            cell = ws.cell(row=r, column=c)
            value = cell.value
            # 标记目标单元格
            if r == row and c == col:
                row_data.append(f'[{value}]')
            else:
                row_data.append(value)
        print(f'行{r}: {row_data}')
print()

# 查找所有工作表
print('=== 所有工作表 ===')
for sheet_name in wb.sheetnames:
    print(f'工作表: {sheet_name}')
print()

# 分析每个工作表
for sheet_name in wb.sheetnames:
    print(f'=== 工作表 {sheet_name} 分析 ===')
    ws = wb[sheet_name]
    print(f'行数: {ws.max_row}')
    print(f'列数: {ws.max_column}')
    
    # 查看该工作表的前5行数据
    print('前5行数据:')
    for row in range(1, min(6, ws.max_row + 1)):
        row_data = []
        for col in range(1, min(10, ws.max_column + 1)):
            cell = ws.cell(row=row, column=col)
            row_data.append(cell.value)
        print(f'行{row}: {row_data}')
    print()

# 重新加载文件以获取公式
wb_formula = openpyxl.load_workbook('外部因素对煤耗的计算表1222.xlsx', data_only=False)

print('=== 所有工作表的公式 ===')
for sheet_name in wb_formula.sheetnames:
    ws = wb_formula[sheet_name]
    print(f'\n工作表: {sheet_name}')
    
    formula_cells = []
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if cell.data_type == 'f':
                formula_cells.append({
                    'cell': get_column_letter(col) + str(row),
                    'formula': cell.value
                })
    
    print(f'公式数量: {len(formula_cells)}')
    for cell_info in formula_cells:
        print(f'  {cell_info["cell"]}: {cell_info["formula"]}')