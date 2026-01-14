import openpyxl
from openpyxl.utils import get_column_letter

# 加载Excel文件
wb = openpyxl.load_workbook('外部因素对煤耗的计算表1222.xlsx', data_only=True)

print('=== 提取Excel中的关键参数 ===\n')

# 查找所有工作表
for sheet_name in wb.sheetnames:
    print(f'工作表: {sheet_name}')
    print('-' * 50)
    
    ws = wb[sheet_name]
    
    # 查找包含关键字的单元格
    keywords = ['基准', '负荷', '温度', '压力', '湿度', '发热量', '效率', '煤耗', '供热']
    
    for row in range(1, min(ws.max_row + 1, 60)):  # 查看前60行
        for col in range(1, min(ws.max_column + 1, 40)):  # 查看前40列
            cell = ws.cell(row=row, column=col)
            if cell.value:
                cell_value = str(cell.value)
                for keyword in keywords:
                    if keyword in cell_value:
                        # 查找该单元格周围的数据
                        print(f'行{row}列{get_column_letter(col)}: {cell.value}')
                        
                        # 查找右侧和下方的数据
                        for offset in range(1, 6):
                            right_cell = ws.cell(row=row, column=col+offset)
                            if right_cell.value and isinstance(right_cell.value, (int, float)):
                                print(f'  -> 右侧{offset}列: {right_cell.value}')
                            
                            down_cell = ws.cell(row=row+offset, column=col)
                            if down_cell.value and isinstance(down_cell.value, (int, float)):
                                print(f'  -> 下方{offset}行: {down_cell.value}')
                        print()
                        break
    print()

# 查看前30行的所有数据
print('\n=== 前30行数据概览 ===')
ws = wb.active
for row in range(1, min(31, ws.max_row + 1)):
    row_data = []
    for col in range(1, min(21, ws.max_column + 1)):  # 前20列
        cell = ws.cell(row=row, column=col)
        if cell.value is not None:
            row_data.append(f'{get_column_letter(col)}:{cell.value}')
    if row_data:
        print(f'行{row}: {", ".join(row_data)}')
