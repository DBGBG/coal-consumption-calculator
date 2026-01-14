import openpyxl
from openpyxl.utils import get_column_letter

# 加载Excel文件
wb = openpyxl.load_workbook('外部因素对煤耗的计算表1222.xlsx', data_only=False)
ws = wb.active

print('=== Excel文件分析报告 ===')
print(f'工作表名称: {ws.title}')
print(f'行数: {ws.max_row}')
print(f'列数: {ws.max_column}')
print()

# 读取表头
print('=== 表头信息 ===')
header_row = 1
headers = []
for col in range(1, ws.max_column + 1):
    cell = ws.cell(row=header_row, column=col)
    headers.append(cell.value)
    print(f'列{get_column_letter(col)}: {cell.value}')
print()

# 读取前10行数据
print('=== 前10行数据 ===')
for row in range(1, min(11, ws.max_row + 1)):
    row_data = []
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        row_data.append(cell.value)
    print(f'行{row}: {row_data}')
print()

# 查找所有包含公式的单元格
print('=== 公式单元格分析 ===')
formula_cells = []
for row in range(1, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        if cell.data_type == 'f':
            formula_cells.append({
                'cell': get_column_letter(col) + str(row),
                'formula': cell.value,
                'row': row,
                'col': col
            })

print(f'共找到 {len(formula_cells)} 个公式单元格')
print()

# 按行分组显示公式
print('=== 按行显示公式 ===')
for cell_info in formula_cells:
    cell_address = cell_info['cell']
    formula = cell_info['formula']
    row = cell_info['row']
    col = cell_info['col']
    
    # 获取单元格的列标题
    col_header = headers[col - 1] if col - 1 < len(headers) else f'列{get_column_letter(col)}'
    
    print(f'单元格 {cell_address} (行{row}, {col_header}):')
    print(f'  公式: {formula}')
    print()

# 分析可能的计算逻辑
print('=== 计算逻辑分析 ===')
print('基于公式单元格的分布，分析可能的计算逻辑...')

# 检查是否有明显的计算列
print()
print('=== 数据类型分析 ===')
for col in range(1, ws.max_column + 1):
    col_letter = get_column_letter(col)
    col_header = headers[col - 1] if col - 1 < len(headers) else f'列{col_letter}'
    
    # 检查该列的数据类型
    numeric_count = 0
    formula_count = 0
    string_count = 0
    
    for row in range(2, ws.max_row + 1):  # 从第二行开始
        cell = ws.cell(row=row, column=col)
        if cell.data_type == 'f':
            formula_count += 1
        elif cell.data_type == 'n':
            numeric_count += 1
        elif cell.data_type == 's':
            string_count += 1
    
    print(f'{col_header} ({col_letter}列):')
    print(f'  数值单元格: {numeric_count}')
    print(f'  公式单元格: {formula_count}')
    print(f'  文本单元格: {string_count}')
    print()