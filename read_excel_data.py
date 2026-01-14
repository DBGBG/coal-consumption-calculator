import openpyxl
from openpyxl.utils import get_column_letter

# 加载Excel文件
wb = openpyxl.load_workbook('外部因素对煤耗的计算表1222.xlsx', data_only=True)

print('=== Excel文件所有工作表 ===\n')
for sheet_name in wb.sheetnames:
    print(f'工作表: {sheet_name}')

print('\n=== 主工作表前60行数据 ===\n')
ws = wb.active

# 查看前60行的所有数据
for row in range(1, min(61, ws.max_row + 1)):
    row_data = []
    for col in range(1, min(41, ws.max_column + 1)):  # 前40列
        cell = ws.cell(row=row, column=col)
        if cell.value is not None:
            cell_str = str(cell.value)
            if len(cell_str) > 15:
                cell_str = cell_str[:15] + "..."
            row_data.append(f'{get_column_letter(col)}:{cell_str}')
    if row_data:
        print(f'行{row}: {", ".join(row_data)}')
